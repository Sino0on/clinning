from rest_framework import viewsets
from .models import *
from .serializers import *
from rest_framework import generics

class ItemViewSet(viewsets.ModelViewSet):
    queryset = Item.objects.all()
    serializer_class = ItemSerializer


import openpyxl
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework import status
from .models import Item
from .serializers import FileUploadSerializer
from .forms import ExcelUploadForm

class ItemImportView(APIView):
    parser_classes = [MultiPartParser]  # Важно для приема файлов

    def post(self, request, format=None):
        serializer = FileUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            # Список для сбора ошибок (если какие-то строки битые)
            errors = []
            
            # Используем атомарную транзакцию: если скрипт упадет, база откатится
            with transaction.atomic():
                # min_row=2 чтобы пропустить заголовок. values_only=True сразу дает значения
                for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    
                    # Проверка на пустую строку (Excel часто грешит "фантомными" строками)
                    if not row or all(cell is None for cell in row):
                        continue
                    
                    # Безопасное извлечение данных (ожидаем 3 колонки)
                    # A -> key, B -> value, C -> value_ru
                    try:
                        key = str(row[0]).strip() if row[0] is not None else None
                        value = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                        value_ru = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

                        if not key:
                            errors.append(f"Строка {index}: Отсутствует ключ (Key)")
                            continue

                        # Логика Upsert (Update or Insert)
                        obj, created = Item.objects.update_or_create(
                            key=key,
                            defaults={
                                'value': value,
                                'value_ru': value_ru
                            }
                        )
                    except Exception as e:
                        # Ловим неожиданные ошибки парсинга
                        errors.append(f"Строка {index}: Ошибка обработки - {str(e)}")

            if errors:
                return Response({
                    "status": "warning",
                    "message": "Импорт завершен с ошибками",
                    "errors": errors
                }, status=status.HTTP_207_MULTI_STATUS)

            return Response({"status": "success", "message": "Данные успешно обновлены"}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from .models import Item
from .forms import ExcelUploadForm

def upload_items_view(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            
            try:
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                errors = []
                created_count = 0
                updated_count = 0

                # Атомарность всё ещё важна
                with transaction.atomic():
                    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(cell is None for cell in row):
                            continue
                        
                        try:
                            # Безопасное получение данных
                            key = str(row[0]).strip() if row[0] is not None else None
                            value = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                            value_ru = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

                            if not key:
                                errors.append(f"Строка {index}: Пропущен key")
                                continue

                            obj, created = Item.objects.update_or_create(
                                key=key,
                                defaults={
                                    'value': value,
                                    'value_ru': value_ru
                                }
                            )
                            
                            if created:
                                created_count += 1
                            else:
                                updated_count += 1

                        except Exception as e:
                            errors.append(f"Строка {index}: {str(e)}")
                
                # Логика отображения результатов
                if errors:
                    # Если были ошибки, показываем их, но транзакция могла пройти частично 
                    # (если ошибка была отловлена внутри цикла, transaction.atomic НЕ откатится, 
                    # если ты не вызовешь raise внутри).
                    # В текущем коде errors ловятся, значит валидные строки сохранятся.
                    messages.warning(request, f"Обработано с ошибками. Создано: {created_count}, Обновлено: {updated_count}.")
                    for err in errors[:10]: # Показываем только первые 10, чтобы не порвать шаблон
                        messages.error(request, err)
                    if len(errors) > 10:
                        messages.error(request, f"...и еще {len(errors) - 10} ошибок.")
                else:
                    messages.success(request, f"Успех! Создано: {created_count}, Обновлено: {updated_count}.")
                
                return redirect('upload_items') # Redirect на ту же страницу (защита от F5)

            except Exception as e:
                messages.error(request, f"Критическая ошибка файла: {str(e)}")
                # Здесь redirect не обязателен, можно вернуть форму с ошибкой
    else:
        form = ExcelUploadForm()

    return render(request, 'upload.html', {'form': form})



class ItemsListView(generics.ListAPIView):
    model = Item
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    


class PartnersListView(generics.ListAPIView):
    model = Partners
    queryset = Partners.objects.all()
    serializer_class = PartnersSerializer


class FeedbackListView(generics.ListAPIView):
    model = Feedback
    queryset = Feedback.objects.all()
    serializer_class = FeedbackSerializer
    

class UseproductListView(generics.ListAPIView):
    model = Useproduct
    queryset = Useproduct.objects.all()
    serializer_class = UseproductSerializer


class BackgroundImageView(generics.ListAPIView):
    model = BackgroundImage
    queryset = BackgroundImage.objects.all()
    serializer_class = BackgroundImageSerializer


class DoiposleListView(generics.ListAPIView):
    model = Doiposle
    queryset = Doiposle.objects.all()
    serializer_class = DoiposleSerializer


